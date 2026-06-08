#!/usr/bin/env python3
"""向日报 xlsx 追加或更新今日行。"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
HEADERS = [
    "日期",
    "星期",
    "每日工作汇报(上午）",
    "每日工作汇报(下午）",
    "明日工作安排",
    "工作注意事项",
]


def _load_config(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _today_cells() -> tuple[str, str]:
    now = datetime.now()
    return now.strftime("%Y.%m.%d"), WEEKDAY_CN[now.weekday()]


def _find_today_row(ws, date_str: str) -> int | None:
    """返回今日日期所在行号（1-based），不存在则 None。"""
    for row in range(2, ws.max_row + 2):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        if str(val).strip() == date_str:
            return row
    return None


def _next_empty_row(ws) -> int:
    for row in range(2, ws.max_row + 50):
        if ws.cell(row=row, column=1).value in (None, ""):
            return row
    return ws.max_row + 1


def _unmerge_columns(ws, row: int, col_start: int, col_end: int) -> None:
    """解除指定行上跨列合并，否则 openpyxl 无法写入 MergedCell。"""
    to_remove: list[str] = []
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_row <= row <= merged.max_row
            and merged.min_col <= col_end
            and merged.max_col >= col_start
        ):
            to_remove.append(str(merged))
    for ref in to_remove:
        ws.unmerge_cells(ref)


def _set_cell(ws, row: int, col: int, value: str) -> None:
    if col in (3, 4):
        _unmerge_columns(ws, row, 3, 4)
    ws.cell(row=row, column=col, value=value)


def _merge_cd_row(ws, row: int) -> None:
    """C 列写满后合并 C:D，与 2026.06.04 等行版式一致。"""
    _unmerge_columns(ws, row, 3, 4)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)


def main() -> int:
    parser = argparse.ArgumentParser(description="更新日报 xlsx")
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--morning", required=True)
    parser.add_argument("--afternoon", required=True)
    parser.add_argument("--tomorrow", default="")
    parser.add_argument("--notes", default="")
    parser.add_argument("--no-overwrite", action="store_true", help="今日已有行时不覆盖")
    parser.add_argument(
        "--merge-cd",
        action="store_true",
        help="写入后合并 C:D（unified 版式，下午列留空）",
    )
    parser.add_argument("--no-merge-cd", action="store_true", help="保持 C、D 分列不合并")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("请先安装: pip3 install openpyxl", file=sys.stderr)
        return 1

    cfg = _load_config(args.config)
    xlsx_path = Path(cfg["xlsx_path"]).expanduser()
    if not xlsx_path.exists():
        print(f"xlsx 不存在: {xlsx_path}", file=sys.stderr)
        return 1

    sheet_name = cfg.get("active_sheet")
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[-1]
        sheet_name = ws.title

    date_str, weekday = _today_cells()
    row = _find_today_row(ws, date_str)
    if row and args.no_overwrite:
        print(f"今日行已存在 (row {row})，已跳过", file=sys.stderr)
        return 2
    if not row:
        row = _next_empty_row(ws)

    values = [date_str, weekday, args.morning, args.afternoon, args.tomorrow, args.notes]
    for col, val in enumerate(values, 1):
        _set_cell(ws, row, col, val)

    if args.merge_cd and not args.no_merge_cd:
        _merge_cd_row(ws, row)

    wb.save(xlsx_path)
    print(f"已写入 sheet={sheet_name} row={row} date={date_str}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
